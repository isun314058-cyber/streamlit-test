# =====================================================
# AI 排樁施工系統 完整版
# =====================================================

import streamlit as st
from PIL import Image, ImageDraw, ImageFont
from pdf2image import convert_from_bytes

import pandas as pd
import numpy as np
import math
import random
import io
import cv2
from streamlit_image_coordinates import streamlit_image_coordinates

# =====================================================
# 頁面設定
# =====================================================

st.set_page_config(
    page_title="AI 排樁施工系統",
    layout="wide"
)

# =====================================================
# 深色模式
# =====================================================

st.markdown("""
<style>

.download-panel{
    position: sticky;
    top: 20px;
}

.stApp{
    background-color:#020617;
    color:white;
}

h1,h2,h3,h4,h5,h6,p,span,label{
    color:white;
}

.stButton>button{
    border-radius:12px;
    height:50px;
    font-size:20px;
    font-weight:bold;
}

.schedule-table{
    width:100%;
    min-width:1400px;
    border-collapse:collapse;
    color:white;
    font-size:16px;
}

.schedule-table th{
    background:#132238;
    color:#ffffff;
    padding:14px;
    text-align:center;
    border:1px solid #2d3b55;
    position:sticky;
    top:0;
    z-index:2;
}

.schedule-table td{
    padding:12px;
    border:1px solid #2d3b55;
    vertical-align:top;
}

.schedule-table tr:nth-child(even){
    background:#0b1730;
}

.schedule-table tr:hover{
    background:#16284a;
}

</style>
""", unsafe_allow_html=True)

# =====================================================
# Title
# =====================================================

st.title("🏗️ AI 排樁施工系統")

# =====================================================
# 功能模式
# =====================================================

mode = st.radio(
    "請選擇功能模式",
    [
        "新建預定進度表",
        "修正當前進度表"
    ],
    horizontal=True,
    key="mode_selector"
)
if "last_mode" not in st.session_state:
    st.session_state.last_mode = mode

if st.session_state.last_mode != mode:

    keys_to_reset = [
    
        "points",
        "last_clicked",
    
        "repair_points",
        "repair_piles",
        "excluded_piles",
        "repair_last_clicked",
        "exclude_last_click",
    
        "repair_canvas_key",
        "repair_current_file",
    
        "schedule_df",
        "repair_schedule_df",
    
        "result_image",
        "original_image",
    
        "pile_positions"
    ]

    for k in keys_to_reset:

        if k in st.session_state:

            if isinstance(st.session_state[k], list):
                st.session_state[k] = []
            else:
                st.session_state[k] = None

    st.session_state.processed = False

    st.session_state.last_mode = mode

    st.rerun()
        
st.markdown("---")

# =====================================================
# Session State
# =====================================================

DEFAULT_STATES = {
    "result_image": None,
    "schedule_df": None,
    "pile_positions": [],
    "processed": False,
    "original_image": None,
    "points": [],
    "last_clicked": None
}

for key, value in DEFAULT_STATES.items():

    if key not in st.session_state:
        st.session_state[key] = value

# =====================================================
# 點位顏色
# =====================================================

POINT_COLORS = [
    ("第一點", "red"),
    ("第二點", "blue"),
    ("第三點", "orange"),
    ("第四點", "lime")
]

COLOR_TEXT = {
    "red": "紅色",
    "blue": "藍色",
    "orange": "橘色",
    "lime": "綠色"
}

# =====================================================
# AI辨識樁位
# =====================================================

def detect_piles(pil_image, roi=None):

    img = np.array(pil_image.convert("RGB"))

    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)

    if roi:

        x1, y1, x2, y2 = roi

        gray = gray[y1:y2, x1:x2]

    gray_blur = cv2.GaussianBlur(
        gray,
        (5, 5),
        1.5
    )
    
    circles = cv2.HoughCircles(
        gray_blur,
        cv2.HOUGH_GRADIENT,
        dp=1.2,
        minDist=25,
        param1=80,
        param2=24,
        minRadius=10,
        maxRadius=18
    )

    positions = []

    if circles is not None:

        circles = np.round(circles[0, :]).astype("int")

        filtered = []

        for (x, y, r) in circles:
            circle_area = math.pi * (r ** 2)
            
            if circle_area < 250:
                continue

            if roi:
                x += x1
                y += y1

            duplicated = False

            for fx, fy, fr in filtered:

                dist = ((x - fx) ** 2 + (y - fy) ** 2) ** 0.5

                if dist < 25:
                    duplicated = True
                    break

            if not duplicated:
                filtered.append((x, y, r))

        filtered = sorted(filtered, key=lambda p: p[1])

        row_tolerance = 40

        grouped_rows = []

        for circle in filtered:

            placed = False

            for row in grouped_rows:

                if abs(circle[1] - row[0][1]) < row_tolerance:
                    row.append(circle)
                    placed = True
                    break

            if not placed:
                grouped_rows.append([circle])

        final_sorted = []

        for row in grouped_rows:

            row = sorted(row, key=lambda p: p[0])

            final_sorted.extend(row)

        # =====================================
        # AI 自動統一樁半徑
        # =====================================
        
        all_radius = [r for (_, _, r) in final_sorted]
        
        valid_radius = [
            r for r in all_radius
            if 8 <= r <= 25
        ]

        if len(valid_radius) == 0:
        
            return []
        
        median_radius = int(np.median(valid_radius))
        row_tolerance = median_radius * 2.2
        
        # 避免極端值
        median_radius = max(10, median_radius)
        median_radius = min(22, median_radius)
        
        positions = []
        
        for (x, y, r) in final_sorted:
        
            positions.append(
                (
                    x,
                    y,
                    median_radius
                )
            )

    return positions
    
@st.cache_data(
    show_spinner=False,
    hash_funcs={Image.Image: id}
)

# =====================================================
# 智慧避鄰排程
# =====================================================

def calculate_distance(p1, p2):

    x1, y1, _ = p1
    x2, y2, _ = p2

    return ((x1 - x2) ** 2 + (y1 - y2) ** 2) ** 0.5


# =====================================================
# 建立鄰樁關係
# =====================================================

def build_neighbor_map(
    pile_positions,
    row_tolerance=40
):

    neighbor_map = {}

    # =========================
    # 依Y座標分列
    # =========================

    pile_data = []

    for idx, (x, y, r) in enumerate(pile_positions):

        pile_data.append({
            "pile": idx + 1,
            "x": x,
            "y": y
        })

    rows = []

    sorted_piles = sorted(
        pile_data,
        key=lambda p: p["y"]
    )

    for pile in sorted_piles:

        found = False

        for row in rows:

            if abs(
                pile["y"] - row[0]["y"]
            ) < row_tolerance:

                row.append(pile)

                found = True

                break

        if not found:

            rows.append([pile])

    # =========================
    # 每列由左到右排序
    # =========================

    for row in rows:

        row.sort(
            key=lambda p: p["x"]
        )

    # =========================
    # 建立上下左右鄰樁
    # =========================

    for row_idx, row in enumerate(rows):

        for col_idx, pile in enumerate(row):

            pile_no = pile["pile"]

            neighbor_map[pile_no] = []

            # 左

            if col_idx > 0:

                neighbor_map[pile_no].append(
                    row[col_idx - 1]["pile"]
                )

            # 右

            if col_idx < len(row) - 1:

                neighbor_map[pile_no].append(
                    row[col_idx + 1]["pile"]
                )

            # 上

            if row_idx > 0:
            
                upper_row = rows[row_idx - 1]
            
                nearest_upper = min(
                    upper_row,
                    key=lambda p:
                    abs(
                        p["x"] - pile["x"]
                    )
                )
            
                if abs(
                    nearest_upper["x"]
                    -
                    pile["x"]
                ) < row_tolerance:
                
                    neighbor_map[pile_no].append(
                        nearest_upper["pile"]
                    )

            # 下

            if row_idx < len(rows) - 1:

                lower_row = rows[row_idx + 1]

                nearest_lower = min(
                    lower_row,
                    key=lambda p:
                    abs(
                        p["x"] - pile["x"]
                    )
                )
                
                if abs(
                    nearest_lower["x"]
                    -
                    pile["x"]
                ) < row_tolerance:
                
                    neighbor_map[pile_no].append(
                        nearest_lower["pile"]
                    )

    # =========================
    # 第二層鄰樁擴充
    # =========================
    
    # for pile_no in neighbor_map:
    
    #     extra_neighbors = []
    
    #     for n in neighbor_map[pile_no]:
    
    #         extra_neighbors.extend(
    #             neighbor_map.get(n, [])
    #         )
    
    #     for n in extra_neighbors:
    
    #         if (
    #             n != pile_no
    #             and
    #             n not in neighbor_map[pile_no]
    #         ):
    
    #             neighbor_map[pile_no].append(n)

    return neighbor_map


def validate_pile_input(edit_df, total_piles):

    import re

    result_df = edit_df.copy()
    
    all_piles = []

    pile_day_map = {}

    error_messages = []

    for idx, row in result_df.iterrows():

        pile_text = str(
            row["施工樁號"]
        ).strip()

        pile_text = pile_text.replace("，", ",")

        # 空白允許
        if pile_text == "":

            result_df.at[idx, "施工數量"] = "0"

            continue

        # 格式檢查
        if not re.fullmatch(
            r"\d+(\s*,\s*\d+)*",
            pile_text
        ):

            result_df.at[idx, "施工數量"] = "輸入錯誤"

            error_messages.append(
                f"{row['施工日']} 輸入格式錯誤"
            )

            continue

        pile_list = [
            int(x.strip())
            for x in pile_text.split(",")
        ]

        # 同一天重複
        if len(pile_list) != len(set(pile_list)):

            result_df.at[idx, "施工數量"] = "重複樁號"

            error_messages.append(
                f"{row['施工日']} 同一天有重複樁號"
            )

            continue

        # 範圍檢查
        out_range = False

        for p in pile_list:

            if p < 1 or p > total_piles:

                result_df.at[idx, "施工數量"] = "樁號超出範圍"

                error_messages.append(
                    f"{row['施工日']} 樁號 {p} 超出範圍"
                )

                out_range = True

                break

        if out_range:

            continue

        result_df.at[idx, "施工數量"] = str(len(pile_list))

        for p in pile_list:

            if p not in pile_day_map:

                pile_day_map[p] = []

            pile_day_map[p].append(
                row["施工日"]
            )

        all_piles.extend(pile_list)

    # =================================
    # 跨天重複檢查
    # =================================

    duplicated_piles = {

        pile

        for pile, days

        in pile_day_map.items()

        if len(days) > 1
    }

    duplicate_detail = {}
    
    for pile, days in pile_day_map.items():
    
        if len(days) > 1:
    
            duplicate_detail[pile] = days

    if duplicated_piles:
        reported = set()
        
        for idx,row in result_df.iterrows():
        
            pile_text = str(
                row["施工樁號"]
            ).strip()
        
            if pile_text == "":
                continue
        
            try:
        
                pile_list = [
                    int(x.strip())
                    for x in pile_text.split(",")
                ]
        
            except:
                continue
        
            dup_list = [
        
                p
        
                for p in pile_list
        
                if p in duplicated_piles
        
            ]
        
            for dup_pile in dup_list:
            
                if dup_pile in reported:
                    continue
            
                reported.add(dup_pile)
            
                error_messages.append(
                    f"樁號 {dup_pile} 重複，出現在："
                    f"{','.join(duplicate_detail[dup_pile])}"
                )
            
    # 最後統一轉字串
    result_df["施工數量"] = (
        result_df["施工數量"]
        .fillna("")
        .astype(str)
    )

    return result_df, error_messages
# =====================================================
# AI 智慧避鄰排程
# =====================================================

def optimize_tail_days(
    schedule,
    neighbor_map,
    daily_count
):

    if len(schedule) < 3:
        return schedule

    max_loop = 20

    for _ in range(max_loop):

        changed = False

        for i in range(len(schedule)-2):

            today = schedule[i]
            last_day = schedule[-1]

            # 最後一天已達目標
            if len(last_day["施工樁號"]) >= daily_count * 0.5:
                return schedule

            # 今日只剩少量不能搬
            if len(today["施工樁號"]) <= 3:
                continue

            move_candidate = None

            if len(today["施工樁號"]) <= daily_count * 0.7:
            
                continue

            for pile in reversed(today["施工樁號"]):

                conflict = False

                for existing in last_day["施工樁號"]:

                    if (
                        pile in neighbor_map.get(existing, [])
                        or
                        existing in neighbor_map.get(pile, [])
                    ):
                        conflict = True
                        break

                if not conflict:
                    move_candidate = pile
                    break

            if move_candidate:

                today["施工樁號"].remove(
                    move_candidate
                )

                last_day["施工樁號"].append(
                    move_candidate
                )

                last_day["施工樁號"].sort()

                changed = True

        if not changed:
            break

    return schedule

def create_schedule(

    pile_positions,
    total_piles,
    daily_count,
    start_date,

    start_no=1,

    cooldown_days=2,

    neighbor_map=None
):

    # 避免 NameError
    future_count = 0
    future_days = 1
    future_avg = 0

    import random
    import pandas as pd

    # =====================================================
    # AI 自動學習最近鄰距離
    # =====================================================

    # ====================================
    # 距離快取
    # ====================================
    
    distance_cache = {}
    
    for i in range(total_piles):
    
        for j in range(i + 1, total_piles):
    
            dist = calculate_distance(
                pile_positions[i],
                pile_positions[j]
            )
    
            distance_cache[(i+1, j+1)] = dist
            distance_cache[(j+1, i+1)] = dist
    
    nearest_distances = []
    
    for i in range(total_piles):
    
        min_dist = 999999
    
        for j in range(total_piles):
    
            if i == j:
                continue
    
            dist = distance_cache.get(
                (i + 1, j + 1),
                999999
            )
    
            if dist < min_dist:
    
                min_dist = dist
    
        nearest_distances.append(
            min_dist
        )
    
    # AI 學習真正樁距
    base_distance = np.median(nearest_distances)

    # =====================================================
    # Delaunay AI 鄰樁判定
    # =====================================================
    
    if neighbor_map is None:
    
        from scipy.spatial import Delaunay
    
        points = np.array([
            [x, y]
            for (x, y, r) in pile_positions
        ])
    
        if len(points) < 3:
    
            return [{
                "施工日": "Day 1",
                "日期": pd.to_datetime(start_date).strftime("%Y-%m-%d"),
                "日期顏色": "#ff6666",
                "施工樁號": list(range(1, total_piles + 1))
            }]
    
        try:
    
            tri = Delaunay(points)
    
        except Exception:
    
            neighbor_map = build_neighbor_map(
                pile_positions
            )
    
        if neighbor_map is None:
    
            neighbor_map = {}
    
            for i in range(len(points)):
    
                neighbor_map[i + 1] = set()
    
            for simplex in tri.simplices:
    
                for i in range(3):
    
                    for j in range(3):
    
                        if i != j:
    
                            p1 = simplex[i] + 1
                            p2 = simplex[j] + 1
    
                            dist = calculate_distance(
                                pile_positions[p1 - 1],
                                pile_positions[p2 - 1]
                            )
    
                            if dist <= base_distance * 1.15:
    
                                neighbor_map[p1].add(p2)
    
            for pile in neighbor_map:
            
                neighbor_map[pile] = sorted(
                    neighbor_map[pile],
                    key=lambda n:
                    calculate_distance(
                        pile_positions[pile-1],
                        pile_positions[n-1]
                    )
                )[:8]
    
            for p in neighbor_map:
    
                neighbors = list(neighbor_map[p])
    
                neighbors = sorted(
                    neighbors,
                    key=lambda n:
                    calculate_distance(
                        pile_positions[p - 1],
                        pile_positions[n - 1]
                    )
                )
    # =====================================================
    # 顏色
    # =====================================================

    colors = []

    for _ in range(100):

        colors.append(
            (
                random.randint(80, 230),
                random.randint(80, 230),
                random.randint(80, 230)
            )
        )

    # =====================================================
    # 初始化
    # =====================================================

    remaining = list(
        range(
            1,
            total_piles + 1
        )
    )
    
    # 預先計算鄰樁數量
    neighbor_score = {}
    
    for p in range(
        1,
        total_piles + 1
    ):
    
        neighbor_score[p] = len(
            neighbor_map.get(p, [])
        )
    
    blocked_until = {}
    
    result = []
    
    day = 1
    TAIL_TRIGGER = daily_count * 1
    # =====================================================
    # 開始排程
    # =====================================================
    loop_guard = 0
    while remaining:
    
        today_piles = []

        # =================================================
        # 第一天第一支固定起始樁
        # =================================================

        if day == 1:

            if start_no in remaining:

                today_piles.append(start_no)

                remaining.remove(start_no)

                blocked_until[start_no] = (
                    day + cooldown_days
                )

                for neighbor in neighbor_map.get(start_no, []):
                    blocked_until[neighbor] = (
                        day + cooldown_days
                    )

        # =================================================
        # AI 智慧選樁
        # =================================================

        tail_mode = False
        
        if len(remaining) <= TAIL_TRIGGER:
            tail_mode = True
            remaining_days = math.ceil(
                len(remaining)
                /
                daily_count
            )
            
            target_tail_count = math.ceil(
                len(remaining)
                /
                remaining_days
            )
        

        today_target = daily_count
        
        if len(remaining) <= TAIL_TRIGGER:
        
            today_target = target_tail_count
        
        while len(today_piles) < today_target:
            
            future_count = max(
                0,
                len(remaining) - 1
            )
            
            safe_daily_count = max(
                1,
                int(daily_count)
            )
            
            future_days = max(
                1,
                math.ceil(
                    future_count / safe_daily_count
                )
            )
            
            future_avg = (
                future_count / future_days
            )
            # =====================================
            # 先過濾還能施工的樁
            # =====================================
        
            candidate_piles = []

            allow_relax = (
                len(remaining)
                <=
                daily_count * 2
            )
            
            for p in remaining:
            
                if p in today_piles:
                    continue
            
                # 前面天數嚴格遵守冷卻
                if not allow_relax:
            
                    if (
                        p in blocked_until
                        and
                        day <= blocked_until[p]
                    ):
                        continue
            
                candidate_piles.append(p)
        
            # 沒候選樁就停止
            if len(candidate_piles) == 0:
                break
        
            # =====================================
            # AI排序
            # =====================================
            #random.shuffle(candidate_piles)          
            sorted_remaining = sorted(
                candidate_piles,
                key=lambda p: neighbor_score[p],
                reverse=True
            )
            
            TOP_K = min(
                60,
                len(sorted_remaining)
            )
            
            sorted_remaining = sorted_remaining[:TOP_K]
            
            random.shuffle(sorted_remaining)
        
            best_score = -999999
        
            best_pile = None
        
            for pile in sorted_remaining:

                # =========================================
                # 冷卻判定
                # =========================================

                if not allow_relax:
                
                    if pile in blocked_until:
                
                        if day <= blocked_until[pile]:
                
                            continue

                # =========================================
                # 十字衝突檢查
                # =========================================

                conflict = False

                for existing in today_piles:

                    if (

                        pile in neighbor_map.get(existing, [])

                        or

                        existing in neighbor_map.get(pile, [])

                    ):

                        conflict = True
                        break

                if conflict:
                    continue

                # =========================================
                # AI 評分
                # =========================================
                
                score = 0
                
                # 鄰居越多越優先
                
                score += len(
                    neighbor_map.get(pile, [])
                ) * 10
                
                # 避免集中同區域
                
                if len(today_piles) > 0:
                
                    min_dist = min(
                
                        distance_cache.get(
                            (pile,p2),
                            999999
                        )
                
                        for p2 in today_piles
                    )
                
                    score += min_dist * 0.2
                
                # =========================================
                # 未來剩餘數量
                # =========================================
            
                score += future_avg * 500
                
                if future_avg < safe_daily_count * 0.8:
                
                    score -= 5000
                
                if (
                    future_count > 0
                    and
                    future_count < safe_daily_count * 0.5
                ):
                
                    score -= 150
                
                # ==================================================
                # 7. 靠近目前群組
                # ==================================================
                
                if len(today_piles) > 0:
                
                    cluster_score = 0
                
                    for existing in today_piles:
                
                        cluster_score += distance_cache.get(
                            (pile, existing),
                            0
                        )
                
                    score -= cluster_score * 0.08
                # =========================================
                # 更新最佳選擇
                # =========================================
                
                if score > best_score:
                
                    best_score = score
                
                    best_pile = pile
                    
            # =============================================
            # 找不到可施工樁
            # =============================================

            if best_pile is None:
                
                
                relaxed_candidates = []
            
                for p in remaining:
            
                    if p in today_piles:
                        continue
            
                    if (
                        p in blocked_until
                        and
                        day <= blocked_until[p]
                    ):
                        continue
            
                    relaxed_candidates.append(p)
            
                if relaxed_candidates:
            
                    best_pile = max(
                        relaxed_candidates,
                        key=lambda p: neighbor_score[p]
                    )
            
                else:
                    break
            # =============================================
            # 加入今日施工
            # =============================================

            today_piles.append(best_pile)

            if best_pile in remaining:

                remaining.remove(best_pile)

            # 立即更新封鎖
            blocked_until[best_pile] = day + cooldown_days
            
            for neighbor in neighbor_map.get(best_pile, []):
            
                blocked_until[neighbor] = (
                    day + cooldown_days
                )
        
        # =================================================
        # 避免卡死
        # =================================================

        if len(today_piles) == 0:
        
            if len(remaining) == 0:
        
                break
        
            first_pile = remaining[0]
        
            today_piles.append(first_pile)
        
            remaining.remove(first_pile)

        # 更新封鎖
        
        if len(remaining) > daily_count * 3:
        
            for pile in today_piles:
        
                blocked_until[pile] = (
                    day + cooldown_days
                )
        
                for neighbor in neighbor_map.get(pile, []):
        
                    blocked_until[neighbor] = (
                        day + cooldown_days
                    )

        # =================================================
        # 日期
        # =================================================

        current_date = (

            pd.to_datetime(start_date)

            + pd.Timedelta(days=day - 1)

        )

        # =================================================
        # 顏色
        # =================================================

        color = colors[(day - 1) % len(colors)]

        hex_color = '#%02x%02x%02x' % color

        # =================================================
        # 儲存結果
        # =================================================

        result.append({

            "施工日": f"Day {day}",

            "日期": current_date.strftime("%Y-%m-%d"),

            "日期顏色": hex_color,

            "施工樁號": sorted(today_piles)

        })

        day += 1
        loop_guard += 1
        
        if loop_guard > total_piles * 2:
        
            st.warning("AI 排程過久，已自動停止")
        
            break

    result = optimize_tail_days(
        result,
        neighbor_map,
        daily_count
    )

    return result

# =====================================================
# 模式：新建預定進度表
# =====================================================

if mode == "新建預定進度表":

    uploaded_file = st.file_uploader(

        "上傳 JPG / PNG / PDF 圖面",

        type=["jpg", "jpeg", "png", "pdf"],

        key="new_schedule"

    )

# =====================================================
# 模式：修正當前進度表
# =====================================================

else:

    uploaded_file = st.file_uploader(

        "上傳目前施工進度圖",

        type=["jpg", "jpeg", "png", "pdf"],

        key="modify_schedule"

    )

    st.info(

        "📌 請上傳目前已編排施工日的圖面"

    )

# =====================================================
# 主流程
# =====================================================
if mode == "新建預定進度表":
        if uploaded_file:
        
            if uploaded_file.type == "application/pdf":
        
                pdf_bytes = uploaded_file.read()
        
                pdf_pages = convert_from_bytes(
                    pdf_bytes,
                    dpi=300
                )
        
                image = pdf_pages[0].convert("RGB")
        
            else:
        
                image = Image.open(uploaded_file).convert("RGB")
        
            st.session_state.original_image = image
        
            MAX_WIDTH = 900
            MAX_HEIGHT = 650
        
            scale_w = MAX_WIDTH / image.width
            scale_h = MAX_HEIGHT / image.height
        
            scale = min(scale_w, scale_h)
        
            display_width = int(image.width * scale)
            display_height = int(image.height * scale)
        
            scale_x = image.width / display_width
            scale_y = image.height / display_height
        
            canvas_bg = image.resize(
                (display_width, display_height)
            )
        
            st.subheader("✏️ 框選施工區域")
        
            st.markdown("""
        
        ✏️ 點選順序
        
        🔴 左上　🔵 左下　🟠 右上　🟢 右下
        """)
        
            left_col, right_col = st.columns([5, 1.3])
        
            preview_canvas = canvas_bg.copy()
        
            draw_preview = ImageDraw.Draw(preview_canvas)
        
            for idx, point in enumerate(st.session_state.points):
        
                px, py = point
        
                label, color = POINT_COLORS[idx]
        
                draw_preview.ellipse(
                    (
                        px - 10,
                        py - 10,
                        px + 10,
                        py + 10
                    ),
                    fill=color,
                    outline="white",
                    width=3
                )
        
                draw_preview.text(
                    (px + 15, py - 15),
                    label,
                    fill=color
                )
        
            roi = None
        
            if len(st.session_state.points) == 4:
        
                xs = [p[0] for p in st.session_state.points]
                ys = [p[1] for p in st.session_state.points]
        
                x1 = min(xs)
                y1 = min(ys)
        
                x2 = max(xs)
                y2 = max(ys)
        
                draw_preview.rectangle(
                    (
                        x1,
                        y1,
                        x2,
                        y2
                    ),
                    outline="lime",
                    width=5
                )
        
                roi = (
                    int(x1 * scale_x),
                    int(y1 * scale_y),
                    int(x2 * scale_x),
                    int(y2 * scale_y)
                )
        
            with left_col:
        
                coords = streamlit_image_coordinates(
                    preview_canvas,
                    key=f"pile_roi_selector_{len(st.session_state.points)}"
                )
        
            if coords is not None:
        
                clicked_point = (
                    coords["x"],
                    coords["y"]
                )
        
                if st.session_state.last_clicked != clicked_point:
        
                    st.session_state.last_clicked = clicked_point
        
                    duplicated = False
        
                    for old_point in st.session_state.points:
        
                        dist = (
                            (clicked_point[0] - old_point[0]) ** 2
                            +
                            (clicked_point[1] - old_point[1]) ** 2
                        ) ** 0.5
        
                        if dist < 10:
                            duplicated = True
                            break
        
                    if (
                        not duplicated
                        and len(st.session_state.points) < 4
                    ):
        
                        st.session_state.points.append(clicked_point)
        
                        st.rerun()
        
            with right_col:
        
                st.subheader("📍 點位資訊")
        
                if len(st.session_state.points) == 0:
        
                    st.info("尚未點選")
        
                else:
        
                    for idx, point in enumerate(st.session_state.points):
        
                        label, color = POINT_COLORS[idx]
        
                        st.markdown(
                            f"""
        {label}
        
        顏色：{COLOR_TEXT[color]}
        """
                        )
        
                st.markdown("---")
        
                if roi:
        
                    st.success("✅ 已完成施工區域")
        
                if st.button("🔄 重新選取"):
                
                    st.session_state.points = []
                
                    st.session_state.last_clicked = None
                
                    st.session_state.pile_positions = []
                
                    st.session_state.schedule_df = None
                
                    st.session_state.result_image = None
                
                    st.session_state.processed = False
                
                    st.rerun()
        
            if roi:
        
                piles = detect_piles(image, roi)
        
                st.session_state.pile_positions = piles

                total_piles = len(piles)
                                        
                st.success(f"✅ AI 辨識到 {total_piles} 支樁體")
        
                result_img = image.copy()
        
                draw = ImageDraw.Draw(result_img)
        
                FONT_NAME = "DejaVuSans.ttf"
                
                try:
                
                    font = ImageFont.truetype(
                        FONT_NAME,
                        18
                    )
                
                    st.success("AI辨識結果字體載入成功")
                
                except Exception as e:
                
                    st.error(f"AI辨識結果字體失敗: {e}")
                
                    font = ImageFont.load_default()
        
                for idx, (x, y, r) in enumerate(piles):
        
                    pile_no = idx + 1
        
                    draw.ellipse(
                        (
                            x - r,
                            y - r,
                            x + r,
                            y + r
                        ),
                        outline="red",
                        width=4
                    )
        
                    pile_text = str(pile_no)

                    pile_bbox = draw.textbbox(
                        (0, 0),
                        pile_text,
                        font=font
                    )
                    
                    pile_width = pile_bbox[2] - pile_bbox[0]
                    
                    pile_x = x - (pile_width // 2) + 1
                    
                    draw.text(
                        (
                            pile_x,
                            y - 38
                        ),
                        pile_text,
                        fill="red",
                        font=font
                    )
        
                # =====================================================
                # AI辨識結果 + 施工條件
                # =====================================================
        
                left_area, right_area = st.columns([3, 1.2])
        
                with left_area:
        
                    st.subheader("🔍 AI辨識結果")
        
                    st.image(
                        result_img,
                        width=900
                    )
        
                with right_area:
        
                    st.subheader("📅 施工條件")
        
                    start_date = st.date_input("施工開始日期")
        
                    daily_count = st.number_input(
                        "每日施工支數",
                        min_value=1,
                        value=14
                    )
        
                    start_no = st.number_input(
                        "起始樁號",
                        min_value=1,
                        max_value=total_piles,
                        value=1
                    )
        
                    
                    execute = st.button(
                        "🚀 執行排程",
                        use_container_width=True
                    )
        
                if execute:
                    with st.spinner("🤖 AI 正在分析最佳施工排程中，請稍候..."):
                        # ============================
                        # 只計算一次鄰樁
                        # ============================
                        
                        median_radius = np.median(
                            [r for _, _, r in piles]
                        )
                        
                        neighbor_map = build_neighbor_map(
                            piles,
                            row_tolerance=int(median_radius * 3)
                        )
            
                        best_schedule = None
                        
                        best_total_score = -999999
                        
                        backup_schedule = None
                        
                        # AI 多次模擬
                        for sim in range(10):
                                                       
                            schedule = create_schedule(
                            
                                pile_positions=piles,
                            
                                total_piles=total_piles,
                            
                                daily_count=daily_count,
                            
                                start_date=start_date,
                            
                                start_no=start_no,
                            
                                cooldown_days=2,
                            
                                neighbor_map=neighbor_map
                            )
                            if backup_schedule is None:
                                backup_schedule = schedule

                            last_day_count = len(
                                schedule[-1]["施工樁號"]
                            )
                            
                            # =====================================
                            # AI 總體評分
                            # =====================================
                        
                            schedule_score = 0

                            # =====================================
                            # 滿載獎勵
                            # =====================================
                            
                            daily_counts = [
                                len(x["施工樁號"])
                                for x in schedule
                            ]
                            
                            full_days = sum(
                                1
                                for day in schedule[:-3]
                                if len(day["施工樁號"]) >= daily_count
                            )
                            
                            first_days_score = 0
                            
                            for c in daily_counts[:5]:
                            
                                first_days_score -= abs(
                                    daily_count - c
                                ) * 3000
                            
                            schedule_score += first_days_score
                            schedule_score += full_days * 5000
                            
                            # 天數越少越好
                            schedule_score -= len(schedule) * 5000
                        
                            # 最後三天不要太少
                            last_days = schedule[-3:]
                        
                            last_count = sum(
                        
                                len(x["施工樁號"])
                        
                                for x in last_days
                            )
                        
                            schedule_score += last_count * 40
                        
                            # 平均施工量穩定
                            daily_counts = [
                        
                                len(x["施工樁號"])
                        
                                for x in schedule
        
                            ]
                            avg_daily = np.mean(daily_counts)
                                
                            schedule_score += avg_daily * 50
                        
                            variance = np.var(daily_counts)
                        
                            # 波動越小越好
                            schedule_score -= variance * 30
                        
                            # =====================================
                            # 尾盤修復
                            # =====================================
                            
                            tail_days = schedule[-5:]
                            
                            tail_total = sum(
                                len(x["施工樁號"])
                                for x in tail_days
                            )
                            
                            # 如果最後三天太少
                            if tail_total < daily_count * 4:
                            
                                schedule_score -= 200
                            
                            # 最後一天不能太少
                            last_day_count = len(schedule[-1]["施工樁號"])
                            
                            if last_day_count <= 2:
                            
                                schedule_score -= 300

                            # =====================================
                            # 尾盤遞減檢查（加強版）
                            # =====================================
                            
                            tail_counts = [
                            
                                len(x["施工樁號"])
                            
                                for x in schedule[-5:]
                            
                            ]

                            # 最後一天不要太少
                            
                            if tail_counts[-1] <= 2:
                            
                                schedule_score -= 500
                            
                            # 倒數第二天不要比最後一天多太多
                            
                            if len(tail_counts) >= 2:
                            
                                if tail_counts[-2] - tail_counts[-1] > 5:
                            
                                    schedule_score -= 200
                            
                            
                            # =====================================
                            # 更新最佳結果
                            # =====================================

                            daily_counts = [
                            
                                len(x["施工樁號"])
                            
                                for x in schedule
                            
                            ]

                            # =========================
                            # 前面天數不得提前掉量
                            # =========================
                            
                            for count in daily_counts[:-3]:
                            
                                diff = daily_count - count
                            
                                if diff > 0:
                            
                                    schedule_score -= diff * 30000

                            # ======================
                            # 尾盤品質
                            # ======================
                            
                            tail_counts = daily_counts[-5:]
                         
                            # =========================
                            # 禁止尾盤反彈
                            # =========================
                            
                            for i in range(len(tail_counts)-1):
                            
                                if tail_counts[i+1] > tail_counts[i]:
                            
                                    schedule_score -= 20000

                            # =========================
                            # 尾盤不可暴跌
                            # =========================
                            
                            for i in range(len(tail_counts)-1):
                            
                                diff = tail_counts[i] - tail_counts[i+1]
                            
                                if diff > 6:
                                
                                    schedule_score -= 8000
                            
                            # 最後一天
                            
                            last_day = tail_counts[-1]
                            
                            if last_day <= 2:
                            
                                schedule_score -= 5000
                            
                            elif last_day <= 5:
                            
                                schedule_score -= 2500
                            
                            elif last_day <= 8:
                            
                                schedule_score -= 1000

                            # ======================
                            # 遞減檢查
                            # ======================
                            
                            for i in range(len(tail_counts)-1):
                            
                                if tail_counts[i] < tail_counts[i+1]:
                            
                                    schedule_score -= 2000

                            tail_avg = np.mean(tail_counts)
                            
                            schedule_score += tail_avg * 200

                            # ======================
                            # 尾盤平衡度
                            # ======================
                            
                            tail_balance_score = 0
                            
                            for count in tail_counts:
                            
                                tail_balance_score -= abs(
                                    count - tail_avg
                                ) * 300
                            
                            schedule_score += tail_balance_score

                            # =========================
                            # 強制尾盤遞減
                            # =========================
                            
                            tail_ok = True
                            
                            tail_counts = [
                                len(x["施工樁號"])
                                for x in schedule[-5:]
                            ]
                            
                            for i in range(
                                len(tail_counts)-1
                            ):
                            
                                if tail_counts[i+1] > tail_counts[i]:
                            
                                    tail_ok = False
                                    break
                            
                            if not tail_ok:
                            
                                continue
                            
                            if schedule_score > best_total_score:
                            
                                best_total_score = schedule_score
                            
                                best_schedule = schedule
                        
                        # 最終最佳排程                        
                        if best_schedule is None:
                        
                            best_schedule = backup_schedule
                        
                        schedule = best_schedule
        
                    df = pd.DataFrame(schedule)
        
                    st.session_state.schedule_df = df
        
                    LEGEND_WIDTH = 165
        
                    new_width = image.width + LEGEND_WIDTH
                    new_height = image.height
        
                    result_img = Image.new(
                        "RGB",
                        (new_width, new_height),
                        (255, 255, 255)
                    )
        
                    result_img.paste(image, (0, 0))
        
                    draw = ImageDraw.Draw(result_img)
        
                    pile_positions = piles
        
                    try:

                        FONT_NAME = "DejaVuSans.ttf"
                    
                        day_font = ImageFont.truetype(
                            FONT_NAME,
                            14
                        )
                    
                        pile_font = ImageFont.truetype(
                            FONT_NAME,
                            18
                        )
                    
                        legend_font = ImageFont.truetype(
                            FONT_NAME,
                            28
                        )
                    
                        st.success("✅ 字型載入成功")
                    
                    except Exception as e:
                    
                        st.error(f"❌ 字型失敗: {e}")
                    
                        day_font = ImageFont.load_default()
                    
                        pile_font = ImageFont.load_default()
                    
                        legend_font = ImageFont.load_default()
        
                    for i, row in df.iterrows():
        
                        hex_color = row["日期顏色"]
        
                        color = tuple(
                        
                            int(hex_color[i:i+2], 16)
                        
                            for i in (1, 3, 5)
                        )
        
                        day_text = row["施工日"].replace("Day ", "D")
        
                        for pile_no in row["施工樁號"]:
        
                            idx = pile_no - 1
        
                            if idx >= len(pile_positions):
                                continue
        
                            x, y, r = pile_positions[idx]
        
                            rr = int(r * 0.85)
        
                            draw.ellipse(
                                (
                                    x - rr,
                                    y - rr,
                                    x + rr,
                                    y + rr
                                ),
                                fill=color,
                                outline="black",
                                width=1
                            )
        
                            # =====================================
                            # 日期固定在圓下方
                            # =====================================
                            
                            day_bbox = draw.textbbox(
                                (0, 0),
                                day_text,
                                font=day_font
                            )
                            
                            day_width = day_bbox[2] - day_bbox[0]
                            
                            day_x = x - (day_width // 2)
                            
                            # 日期往下移一點
                            day_y = y + 14
                            
                            # =====================================
                            # 畫施工日期 D1 D2
                            # =====================================
                            
                            draw.text(
                                (
                                    day_x,
                                    day_y
                                ),
                                day_text,
                                fill="black",
                                font=day_font,
                                stroke_width=2,
                                stroke_fill="white"
                            )
                            
                                                       
                            # =====================================
                            # 樁號固定在圓正上方
                            # 不受圓大小影響
                            # =====================================
                            pile_text = str(pile_no)
                            
                            pile_bbox = draw.textbbox(
                                (0, 0),
                                pile_text,
                                font=pile_font
                            )
                            
                            pile_width = pile_bbox[2] - pile_bbox[0]
                            pile_height = pile_bbox[3] - pile_bbox[1]
                            
                            pile_x = x - (pile_width // 2)
                            
                            # 自動依字體大小調整高度
                            pile_y = y - r - pile_height - 8
                            
                            draw.text(
                                (
                                    pile_x,
                                    pile_y
                                ),
                                pile_text,
                                fill="black",
                                font=pile_font,
                                stroke_width=2,
                                stroke_fill="white"
                            )
        
                    legend_x = image.width + 28
                    legend_y = 80
        
                    draw.text(
                        (
                            legend_x,
                            legend_y - 35
                        ),
                        "施工日顏色對照表",
                        fill="black",
                        font=legend_font
                    )
        
                    legend_height = (len(df) * 32) + 2
        
                    draw.rectangle(
                        (
                            legend_x - 20,
                            legend_y - 10,
                            legend_x + 125,
                            legend_y + legend_height
                        ),
                        outline="black",
                        width=2
                    )
        
                    for i, row in df.iterrows():
        
                        hex_color = row["日期顏色"]
        
                        color = tuple(
                            int(hex_color[i:i+2], 16)
                            for i in (1, 3, 5)
                        )
        
                        yy = legend_y + (i * 30)
        
                        draw.rectangle(
                            (
                                legend_x,
                                yy,
                                legend_x + 24,
                                yy + 20
                            ),
                            fill=color,
                            outline="black"
                        )
        
                        day_no = row["施工日"].replace("Day ", "D")
                        
                        draw.text(
                            (
                                legend_x + 35,
                                yy
                            ),
                            day_no,
                            fill="black",
                            font=pile_font
                        )
        
                    st.session_state.result_image = result_img
                    st.session_state.processed = True
        
                    st.rerun()
        
            # =====================================================
            # 顯示排程結果
            # =====================================================
            
            if st.session_state.processed:
            
                st.markdown("---")
                
                title_col, excel_col = st.columns([8,2])
                
                with title_col:
                
                    st.subheader("📋 施工排程結果")
                
                with excel_col:
                
                    if st.session_state.schedule_df is not None:
                
                        excel_df = st.session_state.schedule_df.copy()
                
                        excel_df["施工數量"] = excel_df["施工樁號"].apply(len)
                
                        excel_df["施工樁號"] = excel_df["施工樁號"].apply(
                            lambda x: ",".join(map(str, x))
                        )
                
                        excel_df = excel_df[
                            [
                                "施工日",
                                "日期",
                                "日期顏色",
                                "施工數量",
                                "施工樁號"
                            ]
                        ]
                
                        excel_buffer = io.BytesIO()
                
                        from openpyxl.styles import PatternFill
                            
                        with pd.ExcelWriter(
                            excel_buffer,
                            engine="openpyxl"
                        ) as writer:
                        
                            excel_df.to_excel(
                                writer,
                                sheet_name="施工排程",
                                index=False
                            )
                        
                            ws = writer.book["施工排程"]
                        
                            # 日期顏色欄位(C欄)
                            for row in range(2, ws.max_row + 1):
                        
                                hex_color = ws[f"C{row}"].value
                        
                                if hex_color:
                        
                                    fill = PatternFill(
                                        fill_type="solid",
                                        start_color=hex_color.replace("#",""),
                                        end_color=hex_color.replace("#","")
                                    )
                        
                                    ws[f"C{row}"].fill = fill
                        
                                    # 不顯示色碼
                                    ws[f"C{row}"].value = ""
                        today_str = pd.Timestamp.today().strftime("%Y%m%d")
                        st.download_button(
                            "📊下載Excel",
                            excel_buffer.getvalue(),
                            file_name=f"{today_str}_AI排樁施工計畫.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                
                df = st.session_state.schedule_df
            
                if df is not None:
            
                    display_df = df.copy()
            
                    display_df["施工數量"] = display_df["施工樁號"].apply(len)
            
                    display_df["施工樁號"] = display_df["施工樁號"].apply(
                        lambda x: ", ".join(map(str, x))
                    )
                    # 刪除 RGB 欄位
                    if "RGB" in display_df.columns:
            
                        display_df = display_df.drop(columns=["RGB"])
                
                    # 日期顏色改成色塊
                    display_df["日期顏色"] = display_df["日期顏色"].apply(
                        lambda c:
                        f'<div style="background:{c}; width:80px; height:28px; border-radius:6px;"></div>'
                    )
            
                    display_df = display_df[
                        [
                            "施工日",
                            "日期",
                            "日期顏色",
                            "施工數量",
                            "施工樁號"
                        ]
                    ]
                    
                    st.markdown(
                        f"""
                        <div style="
                            width:100%;
                            overflow-x:auto;
                            overflow-y:auto;
                            max-height:650px;
                            border:1px solid #333;
                            border-radius:12px;
                            padding:10px;
                            background:#071225;
                        ">
                            {display_df.to_html(
                                escape=False,
                                index=False,
                                classes="schedule-table"
                            )}
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
            
                   
            
                # =====================================================
                # 排樁施工圖 + 下載圖面
                # =====================================================
            
                left_result, right_download = st.columns(
                    [4,1.2]
                )
                
                with left_result:
                
                    st.subheader("🗺️ 排樁施工圖")
                
                    st.image(
                        st.session_state.result_image,
                        use_container_width=True
                    )
                
                with right_download:
                    st.markdown(
                        '<div class="download-panel">',
                        unsafe_allow_html=True
                    )
                    st.subheader("📥 下載圖面")
                
                    export_type = st.selectbox(
                        "選擇匯出格式",
                        ["PNG","JPG","PDF"]
                    )
                
                    img_buffer = io.BytesIO()
                
                    result_img = st.session_state.result_image
            
                    if export_type == "PNG":
            
                        result_img.save(img_buffer, format="PNG")
            
                        st.download_button(
                            label="下載 PNG 圖面",
                            data=img_buffer.getvalue(),
                            file_name="pile_schedule.png",
                            mime="image/png",
                            use_container_width=True
                        )
                        
                    elif export_type == "JPG":
            
                        rgb_img = result_img.convert("RGB")
            
                        rgb_img.save(img_buffer, format="JPEG")
            
                        st.download_button(
                            label="下載 JPG 圖面",
                            data=img_buffer.getvalue(),
                            file_name="pile_schedule.jpg",
                            mime="image/jpeg",
                            use_container_width=True
                        )
                        st.markdown(
                            '</div>',
                            unsafe_allow_html=True
                        )
            
                    else:
            
                        rgb_img = result_img.convert("RGB")
            
                        rgb_img.save(img_buffer, format="PDF")
            
                        st.download_button(
                            label="下載 PDF 圖面",
                            data=img_buffer.getvalue(),
                            file_name="pile_schedule.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )

                    st.markdown(
                        '</div>',
                        unsafe_allow_html=True
                    )
elif mode == "修正當前進度表":
    # ============================================
    # 初始化修正模式
    # ============================================
    
    if "repair_mode_init" not in st.session_state:

        st.session_state.repair_points = []
    
        st.session_state.repair_last_clicked = None
    
        st.session_state.repair_piles = []
    
        st.session_state.excluded_piles = []
    
        st.session_state.repair_canvas_key = 0
    
        st.session_state.repair_mode_init = True
    
    # ============================================
    # 上傳圖面
    # ============================================
    
    if uploaded_file:
        current_file_name = uploaded_file.name
    
        if (
            "repair_current_file"
            not in st.session_state
        ):
    
            st.session_state.repair_current_file = current_file_name
    
        elif (
            st.session_state.repair_current_file
            != current_file_name
        ):
    
            st.session_state.repair_points = []
    
            st.session_state.repair_last_clicked = None
    
            st.session_state.repair_current_file = current_file_name

        piles = []

        st.markdown("---")

        st.subheader("🛠️ 修正當前施工進度")

        # ============================================
        # 讀取圖面
        # ============================================

        if uploaded_file.type == "application/pdf":

            pdf_bytes = uploaded_file.read()

            pdf_pages = convert_from_bytes(
                pdf_bytes,
                dpi=300
            )

            image = pdf_pages[0].convert("RGB")

        else:

            image = Image.open(
                uploaded_file
            ).convert("RGB")

        # ============================================
        # 顯示圖面
        # ============================================

        display_img = image.copy()

        display_img.thumbnail((900, 650))
        scale_x = image.width / display_img.width
        scale_y = image.height / display_img.height

        draw_img = display_img.copy()

        draw = ImageDraw.Draw(draw_img)

        point_colors = [
            "red",
            "blue",
            "orange",
            "lime"
        ]

        # ============================================
        # 畫點
        # ============================================

        for idx, point in enumerate(
            st.session_state.repair_points
        ):

            x, y = point

            draw.ellipse(
                (x-8, y-8, x+8, y+8),
                fill=point_colors[idx]
            )

        # ============================================
        # 畫框
        # ============================================
        
        if len(st.session_state.repair_points) == 4:
        
            pts = st.session_state.repair_points
        
            xs = [p[0] for p in pts]
            ys = [p[1] for p in pts]
        
            x1 = min(xs)
            y1 = min(ys)
        
            x2 = max(xs)
            y2 = max(ys)
        
            draw.rectangle(
                (
                    x1,
                    y1,
                    x2,
                    y2
                ),
                outline="lime",
                width=5
            )



        # ============================================
        # 左右欄位
        # ============================================
        
        left_col, right_col = st.columns([5, 1.3])
        
        # ============================================
        # 左側圖面
        # ============================================
        
        with left_col:
        
            value = streamlit_image_coordinates(
                draw_img,
                key=f"repair_roi_{st.session_state.repair_canvas_key}"
            )
        
        # ============================================
        # 右側資訊
        # ============================================
        
        with right_col:
        
            st.subheader("📍 點位資訊")
        
            if len(st.session_state.repair_points) == 0:
        
                st.info("尚未點選")
        
            else:
        
                point_names = [
                    ("第一點", "紅色"),
                    ("第二點", "藍色"),
                    ("第三點", "橘色"),
                    ("第四點", "綠色")
                ]
        
                for idx, point in enumerate(
                    st.session_state.repair_points
                ):
        
                    name, color = point_names[idx]
        
                    st.markdown(
                        f"""
        {name}
        
        顏色：{color}
        """
                    )
        
            st.markdown("---")
        
            if len(st.session_state.repair_points) == 4:
        
                st.success("✅ 已完成施工區域")
        
            if st.button("🔄 重新選取"):

                st.session_state.repair_points = []
            
                st.session_state.repair_last_clicked = None
            
                st.session_state.repair_piles = []

                st.session_state.excluded_piles = []
            
                st.session_state.repair_canvas_key += 1
            
                st.rerun()
        
        # ============================================
        # 點擊新增
        # ============================================
        
        if value is not None:
        
            clicked_point = (
                value["x"],
                value["y"]
            )
        
            if (
                st.session_state.repair_last_clicked
                != clicked_point
            ):
        
                st.session_state.repair_last_clicked = clicked_point
        
                duplicated = False
                
                for old_point in st.session_state.repair_points:
                
                    dist = (
                        (clicked_point[0] - old_point[0]) ** 2
                        +
                        (clicked_point[1] - old_point[1]) ** 2
                    ) ** 0.5
                
                    if dist < 10:
                        duplicated = True
                        break
                
                if (
                    not duplicated
                    and len(st.session_state.repair_points) < 4
                ):
                
                    st.session_state.repair_points.append(
                        clicked_point
                    )
                
                    st.rerun()

        # ============================================
        # ROI完成 → AI辨識
        # ============================================

        if len(st.session_state.repair_points) == 4:

            pts = st.session_state.repair_points

            xs = [p[0] for p in pts]
            ys = [p[1] for p in pts]
            
            x1 = min(xs)
            y1 = min(ys)
            
            x2 = max(xs)
            y2 = max(ys)

            roi = (
                int(x1 * scale_x),
                int(y1 * scale_y),
                int(x2 * scale_x),
                int(y2 * scale_y)
            )

            # AI辨識（只執行一次）

            if len(st.session_state.repair_piles) == 0:
            
                piles = detect_piles(
                    image,
                    roi
                )
            
                st.session_state.repair_piles = piles
            
            else:
            
                piles = st.session_state.repair_piles

            total_piles = len(piles)

            full_neighbor_map = build_neighbor_map(
                piles
            )

            st.session_state.repair_total_piles = total_piles

            st.success(
                f"✅ AI辨識到 {total_piles} 支樁體"
            )

            # ========================================
            # 顯示辨識結果
            # ========================================

            result_img = image.copy()

            display_result = result_img.copy()
            
            display_result.thumbnail((900,650))
            
            scale_x = result_img.width / display_result.width
            scale_y = result_img.height / display_result.height

            draw_result = ImageDraw.Draw(display_result)

            try:

                FONT_NAME = "DejaVuSans.ttf"
                
                font = ImageFont.truetype(
                    FONT_NAME,
                    12
                )

            except:

                font = ImageFont.load_default()

            for idx, (x, y, r) in enumerate(piles):
            
                x = int(x / scale_x)
                y = int(y / scale_y)
                r = int(r / scale_x)
            
                pile_no = idx + 1
            
                # 已排除樁
                if pile_no in st.session_state.excluded_piles:
            
                    draw_result.line(
                        (x-r, y-r, x+r, y+r),
                        fill="red",
                        width=5
                    )
            
                    draw_result.line(
                        (x+r, y-r, x-r, y+r),
                        fill="red",
                        width=5
                    )
            
                    continue
            
                # 畫圓
                draw_result.ellipse(
                    (
                        x-r,
                        y-r,
                        x+r,
                        y+r
                    ),
                    outline="red",
                    width=2
                )
            
                # ===================
                # 新增樁號
                # ===================
            
                pile_text = str(pile_no)
                
                pile_bbox = draw_result.textbbox(
                    (0,0),
                    pile_text,
                    font=font
                )
                
                pile_width = pile_bbox[2] - pile_bbox[0]
                pile_height = pile_bbox[3] - pile_bbox[1]
                
                draw_result.text(
                    (
                        x - pile_width // 2,
                        y - 22
                    ),
                    pile_text,
                    fill="red",
                    font=font
                )
                
            left_result, right_result = st.columns([2.2, 1])
            
            with left_result:
            
                clicked = streamlit_image_coordinates(
                    display_result,
                    key="exclude_pile_click"
                )
            
                if "exclude_last_click" not in st.session_state:
            
                    st.session_state.exclude_last_click = None
            
                if clicked is not None:
            
                    current_click = (
                        clicked["x"],
                        clicked["y"]
                    )
            
                    if (
                        st.session_state.exclude_last_click
                        != current_click
                    ):
            
                        st.session_state.exclude_last_click = current_click
            
                        click_x = clicked["x"] * scale_x
                        click_y = clicked["y"] * scale_y
            
                        nearest_pile = None
                        nearest_dist = 999999
            
                        for idx, (x, y, r) in enumerate(piles):
            
                            dist = (
                                (click_x - x) ** 2
                                +
                                (click_y - y) ** 2
                            ) ** 0.5
            
                            if dist < nearest_dist:
            
                                nearest_dist = dist
                                nearest_pile = idx + 1
            
                        if nearest_dist < 30:
            
                            # 已排除 → 取消排除
                            if (
                                nearest_pile
                                in st.session_state.excluded_piles
                            ):
            
                                st.session_state.excluded_piles.remove(
                                    nearest_pile
                                )
            
                            # 未排除 → 排除
                            else:
            
                                st.session_state.excluded_piles.append(
                                    nearest_pile
                                )
            
                            st.rerun()
            
                
            
            with right_result:
            
                st.subheader("📊 AI辨識結果")

        # ============================================
        # 有辨識到樁體才往下
        # ============================================

        if len(st.session_state.repair_piles) > 0:
        
            st.markdown("---")
            st.subheader("📊 上傳原施工排程 Excel")
        
            excel_file = st.file_uploader(
                "上傳施工排程Excel",
                type=["xlsx"],
                key="repair_excel"
            )
        
            if excel_file:
        
                try:
        
                    original_df = pd.read_excel(excel_file)
        
                    required_cols = [
                        "施工日",
                        "日期",
                        "日期顏色",
                        "施工數量",
                        "施工樁號"
                    ]
        
                    missing_cols = [
                        c
                        for c in required_cols
                        if c not in original_df.columns
                    ]
        
                    if missing_cols:
        
                        st.error(
                            f"缺少欄位：{','.join(missing_cols)}"
                        )
        
                    else:
                                                     
                        st.markdown("---")
                        st.write("✏️ 請於下方修改施工樁號")
                        
                        editor_df = original_df.drop(
                            columns=["日期顏色"]
                        )

                        editor_df["施工數量"] = (
                            editor_df["施工數量"]
                            .astype(str)
                        )
                        if "repair_edit_df" not in st.session_state:
                        
                            st.session_state.repair_edit_df = editor_df.copy()

                        if "repair_excel_name" not in st.session_state:
                            st.session_state.repair_excel_name = ""
                        
                        if (
                            st.session_state.repair_excel_name
                            != excel_file.name
                        ):
                        
                            st.session_state.repair_excel_name = excel_file.name
                        
                            st.session_state.repair_edit_df = editor_df.copy()
                            
                        edited_df = st.data_editor(
                        
                            st.session_state.repair_edit_df,
                        
                            use_container_width=True,
                        
                            height=500,
                        
                            hide_index=True,
                        
                            disabled=[
                                "施工日",
                                "日期",
                                "施工數量"
                            ],
                        
                            key="repair_editor"
                        )
                        
                        # 只有結果不同才更新
                        validated_df, error_messages = validate_pile_input(
                            edited_df,
                            st.session_state.repair_total_piles
                        )
                        
                        if not validated_df.equals(
                            st.session_state.repair_edit_df
                        ):
                            st.session_state.repair_edit_df = validated_df.copy()
                            st.rerun()
                        
                        st.markdown("### 🔍 驗證結果")
                        
                        if error_messages:
                        
                            st.error(
                                "\n".join(error_messages)
                            )
                        
                        else:
                        
                            if len(error_messages) == 0:
                            
                                st.success("✅ 更改完成")

                        # ============================================
                        # 重新排程
                        # ============================================
                        
                        if "repair_edit_df" in st.session_state:
                        
                            st.markdown("---")
                        
                            st.subheader("🚀 重新排程")
                        
                            daily_count = st.number_input(
                                "每日施工支數",
                                min_value=1,
                                value=14,
                                key="repair_daily_count"
                            )
                        
                            if st.button(
                                "🚀重新產出排程",
                                use_container_width=True
                            ):
                        
                                edit_df = st.session_state.repair_edit_df
                                
                                completed_piles = []

                                first_empty_index = None
                                
                                for idx,row in edit_df.iterrows():
                                
                                    pile_text = str(
                                        row["施工樁號"]
                                    ).strip()
                                
                                    if (
                                        pile_text == ""
                                        or pile_text.lower() == "nan"
                                    ):
                                
                                        first_empty_index = idx
                                
                                        break

                                # 找不到空白列

                                if first_empty_index is None:
                                
                                    st.error("全部施工日都有樁號，沒有可續排的施工日")
                                
                                    st.stop()

                                start_day_no = first_empty_index + 1
                                
                                
                                # 自動抓續排開始日期
                                
                                start_date = pd.to_datetime(
                                
                                    edit_df.iloc[first_empty_index]["日期"]
                                
                                )
                                
                                for idx, row in edit_df.iterrows():
                                    if idx >= first_empty_index:
                                        break
                                    pile_text = str(
                                        row["施工樁號"]
                                    ).strip()
                                
                                    if (
                                        pile_text != ""
                                        and pile_text.lower() != "nan"
                                    ):
                                
                                        pile_list = [
                                
                                            int(x.strip())
                                
                                            for x in pile_text.split(",")
                                
                                            if x.strip().isdigit()
                                
                                        ]
                                
                                        completed_piles.extend(
                                            pile_list
                                        )

                                all_piles = set(
                                    range(
                                        1,
                                        st.session_state.repair_total_piles + 1
                                    )
                                )
                                
                                remaining_piles = sorted(
                                    list(
                                        all_piles
                                        -
                                        set(completed_piles)
                                    )
                                )
                                remaining_piles = [
                                    p
                                    for p in remaining_piles
                                    if p not in st.session_state.excluded_piles
                                ]

                                # ==================================
                                # 建立剩餘樁體座標
                                # ==================================
                                
                                remaining_positions = [
                                    piles[p-1]
                                    for p in remaining_piles
                                ]

                                # ==================================
                                # 建立樁號對照表
                                # ==================================
                                
                                pile_mapping = {}
                                
                                for new_no, old_no in enumerate(remaining_piles, start=1):
                                
                                    pile_mapping[new_no] = old_no

                                neighbor_map = {}
                                
                                reverse_mapping = {}
                                
                                for new_no, old_no in pile_mapping.items():
                                
                                    reverse_mapping[old_no] = new_no
                                
                                
                                for old_pile in remaining_piles:
                                
                                    new_pile = reverse_mapping[old_pile]
                                
                                    neighbor_map[new_pile] = []
                                
                                    for n in full_neighbor_map.get(old_pile, []):
                                
                                        if n in remaining_piles:
                                
                                            neighbor_map[new_pile].append(
                                                reverse_mapping[n]
                                            )

                                # ==================================
                                # AI續排
                                # ==================================
                                
                                best_schedule = None
                                
                                best_total_score = -999999

                                backup_schedule = None
                                
                                for sim in range(10):
                                
                                    temp_schedule = create_schedule(
                                    
                                        pile_positions=remaining_positions,
                                    
                                        total_piles=len(remaining_positions),
                                    
                                        daily_count=daily_count,
                                    
                                        start_date=start_date,
                                    
                                        start_no=random.randint(
                                            1,
                                            len(remaining_positions)
                                        ),
                                    
                                        cooldown_days=2,
                                    
                                        neighbor_map=neighbor_map
                                    )

                                    temp_schedule = optimize_tail_days(
                                        temp_schedule,
                                        neighbor_map,
                                        daily_count
                                    )

                                    if backup_schedule is None:
                                    
                                        backup_schedule = temp_schedule

                                    daily_counts = [
                                    
                                        len(x["施工樁號"])
                                    
                                        for x in temp_schedule
                                    
                                    ]
                                
                                    schedule_score = 0
                                    
                                    daily_counts = [
                                        len(x["施工樁號"])
                                        for x in temp_schedule
                                    ]
                                    
                                    full_days = sum(
                                        1
                                        for day in temp_schedule[:-3]
                                        if len(day["施工樁號"]) >= daily_count
                                    )
                                    
                                    first_days_score = 0
                                    
                                    for c in daily_counts[:5]:
                                    
                                        first_days_score -= abs(
                                            daily_count - c
                                        ) * 3000
                                    
                                    schedule_score += first_days_score
                                    schedule_score += full_days * 5000
                                    
                                    schedule_score -= len(temp_schedule) * 5000

                                    last_days = temp_schedule[-3:]
                                    
                                    last_count = sum(
                                        len(x["施工樁號"])
                                        for x in last_days
                                    )
                                    
                                    schedule_score += last_count * 40
                                    
                                    avg_daily = np.mean(daily_counts)
                                    
                                    schedule_score += avg_daily * 50
                                    
                                    variance = np.var(daily_counts)
                                    
                                    schedule_score -= variance * 30

                                    tail_days = temp_schedule[-5:]
                                    
                                    tail_total = sum(
                                        len(x["施工樁號"])
                                        for x in tail_days
                                    )
                                    
                                    if tail_total < daily_count * 4:
                                        schedule_score -= 200
                                    
                                    last_day_count = len(
                                        temp_schedule[-1]["施工樁號"]
                                    )
                                    
                                    if last_day_count <= 2:
                                        schedule_score -= 300

                                    tail_counts = [
                                        len(x["施工樁號"])
                                        for x in temp_schedule[-5:]
                                    ]

                                    for count in daily_counts[:-3]:
                                    
                                        diff = daily_count - count
                                    
                                        if diff > 0:
                                    
                                            schedule_score -= diff * 30000
                                    
                                    
                                    for i in range(len(tail_counts)-1):
                                    
                                        if tail_counts[i+1] > tail_counts[i]:
                                    
                                            schedule_score -= 20000
                                    
                                    
                                    for i in range(len(tail_counts)-1):
                                    
                                        diff = tail_counts[i] - tail_counts[i+1]
                                    
                                        if diff > 6:
                                    
                                            schedule_score -= 8000
                                    
                                    
                                    last_day = tail_counts[-1]
                                    
                                    if last_day <= 2:
                                    
                                        schedule_score -= 5000
                                    
                                    elif last_day <= 5:
                                    
                                        schedule_score -= 2500
                                    
                                    elif last_day <= 8:
                                    
                                        schedule_score -= 1000
                                    
                                    
                                    tail_avg = np.mean(tail_counts)
                                    
                                    schedule_score += tail_avg * 200
                                    
                                    
                                    tail_balance_score = 0
                                    
                                    for count in tail_counts:
                                    
                                        tail_balance_score -= abs(
                                            count - tail_avg
                                        ) * 300
                                    
                                    schedule_score += tail_balance_score
                                    
                                    
                                    tail_ok = True
                                    
                                    for i in range(len(tail_counts)-1):
                                    
                                        if tail_counts[i+1] > tail_counts[i]:
                                    
                                            tail_ok = False
                                    
                                            break
                                    
                                    if not tail_ok:
                                    
                                        continue
                                    
                                    
                                    if schedule_score > best_total_score:
                                    
                                        best_total_score = schedule_score
                                    
                                        best_schedule = temp_schedule

                                # ==================================
                                # 如果10次模擬都失敗
                                # ==================================
                                
                                if best_schedule is None:
                                
                                    best_schedule = backup_schedule
                                
                                # ==================================
                                # 轉回原始樁號
                                # ==================================
                                
                                new_schedule = best_schedule
                                
                                for day in new_schedule:
                                
                                    day["施工樁號"] = [
                                
                                        pile_mapping[p]
                                
                                        for p in day["施工樁號"]
                                
                                    ]

                                    

                                # ==================================
                                # 回填到原排程
                                # ==================================
                                
                                new_df = original_df.copy()

                                new_df["施工樁號"] = edit_df["施工樁號"]
                                new_df["施工數量"] = edit_df["施工數量"]
                                
                                for i, day_data in enumerate(new_schedule):
                                
                                    target_row = first_empty_index + i
                                
                                    if target_row >= len(new_df):
                                    
                                        st.warning(
                                            "Excel施工日不足，部分續排資料未寫入"
                                        )
                                    
                                        break
                                
                                    pile_text = ",".join(
                                        map(str, day_data["施工樁號"])
                                    )
                                
                                    new_df.at[target_row, "施工樁號"] = pile_text

                                    new_df.at[target_row, "日期顏色"] = day_data["日期顏色"]
                                
                                    new_df.at[target_row, "施工數量"] = str(
                                        len(day_data["施工樁號"])
                                    )

                                new_df = new_df[
                                    new_df["施工數量"].astype(str) != "0"
                                ].reset_index(drop=True)

                                st.session_state.repair_schedule_df = new_df
                                
                                st.success("✅ AI續排完成")
                                
                                repair_df = st.session_state.repair_schedule_df.copy()
                                
                                repair_df["施工數量"] = repair_df["施工樁號"].apply(
                                    lambda x:
                                    len([
                                        p for p in str(x).split(",")
                                        if p.strip()
                                    ])
                                    if pd.notna(x)
                                    else 0
                                )

                                # 刪除施工數量為0的列
                                repair_df = repair_df[
                                    repair_df["施工數量"] > 0
                                ].reset_index(drop=True)
                                
                                st.dataframe(
                                    repair_df,
                                    use_container_width=True,
                                    hide_index=True
                                )
                                
                                # =====================================
                                # 產生續排施工圖
                                # =====================================
                                
                                repair_result_img = image.copy()
                                
                                draw = ImageDraw.Draw(repair_result_img)

                                try:
                                
                                    pile_font = ImageFont.truetype(
                                        "DejaVuSans.ttf",
                                        18
                                    )

                                    day_font = ImageFont.truetype(
                                        "DejaVuSans.ttf",
                                        14
                                    )
                                
                                except:
                                
                                    pile_font = ImageFont.load_default()

                                LEGEND_WIDTH = 165
                                
                                new_width = image.width + LEGEND_WIDTH
                                
                                repair_result_img = Image.new(
                                    "RGB",
                                    (new_width, image.height),
                                    (255,255,255)
                                )
                                
                                repair_result_img.paste(
                                    image,
                                    (0,0)
                                )
                                
                                draw = ImageDraw.Draw(
                                    repair_result_img
                                )

                                for pile_no in completed_piles:
                                
                                    idx = pile_no - 1
                                
                                    x,y,r = piles[idx]
                                
                                    draw.ellipse(
                                        (
                                            x-r,
                                            y-r,
                                            x+r,
                                            y+r
                                        ),
                                        fill=(180,180,180),
                                        outline="black",
                                        width=2
                                    )

                                for day_idx,row in enumerate(new_schedule):
                                
                                    hex_color = row["日期顏色"]
                                
                                    color = tuple(
                                        int(hex_color[i:i+2],16)
                                        for i in (1,3,5)
                                    )
                                
                                    for pile_no in row["施工樁號"]:
                                
                                        idx = pile_no - 1
                                
                                        x,y,r = piles[idx]
                                
                                        draw.ellipse(
                                            (
                                                x-r,
                                                y-r,
                                                x+r,
                                                y+r
                                            ),
                                            fill=color,
                                            outline="black",
                                            width=2
                                        )
                                
                                        pile_text = str(pile_no)
                                
                                        pile_bbox = draw.textbbox(
                                            (0, 0),
                                            pile_text,
                                            font=pile_font
                                        )
                                
                                        pile_width = pile_bbox[2] - pile_bbox[0]
                                
                                        pile_x = x - (pile_width // 2)
                                
                                        draw.text(
                                            (
                                                pile_x,
                                                y - r - 25
                                            ),
                                            pile_text,
                                            fill="black",
                                            font=pile_font
                                        )   

                                        day_text = f"D{start_day_no + day_idx}"
                                        
                                        day_bbox = draw.textbbox(
                                            (0,0),
                                            day_text,
                                            font=pile_font
                                        )
                                        
                                        day_width = day_bbox[2] - day_bbox[0]
                                        
                                        draw.text(
                                            (
                                                x - day_width // 2,
                                                y + r + 8
                                            ),
                                            day_text,
                                            fill="black",
                                            font=day_font,
                                            stroke_width=2,
                                            stroke_fill="white"
                                        )

                                legend_x = image.width + 25
                                
                                legend_y = 80

                                draw.text(
                                    (
                                        legend_x,
                                        legend_y - 35
                                    ),
                                    "施工日圖例",
                                    fill="black",
                                    font=pile_font
                                )

                                for day_idx,row in enumerate(new_schedule):
                                
                                    hex_color = row["日期顏色"]
                                
                                    color = tuple(
                                        int(hex_color[i:i+2],16)
                                        for i in (1,3,5)
                                    )
                                
                                    yy = legend_y + day_idx * 30
                                
                                    draw.rectangle(
                                        (
                                            legend_x,
                                            yy,
                                            legend_x+22,
                                            yy+22
                                        ),
                                        fill=color,
                                        outline="black"
                                    )
                                
                                    draw.text(
                                        (
                                            legend_x+35,
                                            yy
                                        ),
                                        f"D{start_day_no + day_idx}",
                                        fill="black",
                                        font=pile_font
                                    )

                                st.session_state.repair_result_image = repair_result_img
                                
                                st.image(
                                    repair_result_img,
                                    width=900
                                )

                                excel_buffer = io.BytesIO()
                                
                                with pd.ExcelWriter(
                                    excel_buffer,
                                    engine="openpyxl"
                                ) as writer:
                                
                                    repair_df.to_excel(
                                        writer,
                                        index=False
                                    )
                                
                                st.download_button(
                                    "📊下載續排Excel",
                                    excel_buffer.getvalue(),
                                    "repair_schedule.xlsx",
                                    use_container_width=True
                                )

                                png_buffer = io.BytesIO()
                                
                                repair_result_img.save(
                                    png_buffer,
                                    format="PNG"
                                )
                                
                                st.download_button(
                                    "🖼️下載PNG",
                                    png_buffer.getvalue(),
                                    "repair_schedule.png",
                                    use_container_width=True
                                )

                                jpg_buffer = io.BytesIO()
                                
                                repair_result_img.convert("RGB").save(
                                    jpg_buffer,
                                    format="JPEG"
                                )
                                
                                st.download_button(
                                    "🖼️下載JPG",
                                    jpg_buffer.getvalue(),
                                    "repair_schedule.jpg",
                                    use_container_width=True
                                )

                                pdf_buffer = io.BytesIO()
                                
                                repair_result_img.convert("RGB").save(
                                    pdf_buffer,
                                    format="PDF"
                                )
                                
                                st.download_button(
                                    "📄下載PDF",
                                    pdf_buffer.getvalue(),
                                    "repair_schedule.pdf",
                                    use_container_width=True
                                )

                
                except Exception as e:
        
                    st.error(f"Excel讀取失敗：{e}")
